import pandas as pd
import numpy as np
from termcolor import colored
from openpyxl.styles import PatternFill
import sys
import pdb
import os
from pathlib import Path
from PyQt5.uic import loadUiType
from PyQt5.QtGui import QIcon, QPixmap, QColor, QFont
from PyQt5.QtCore import Qt, pyqtSignal, QSize, QObject, QTimer, QThread, QThreadPool, QRect
from PyQt5.uic import loadUi
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QWidget,
    QMessageBox,
    QFileDialog

)
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import socket
import uuid
import jdatetime


ui, _ = loadUiType("AccuntingApp.ui")


############ Notification class ############
class Message(QWidget):
    def __init__(self, title, message, icon, parent=None):
        QWidget.__init__(self, parent)
        self.widget = QWidget(self)
        self.widget.setObjectName('Custom_Widget')
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(self.widget)
        self.layout = QGridLayout(self.widget)

        self.titleLabel = QLabel(title, self)
        self.titleLabel.setStyleSheet(
            "font-size: 18px; font-weight: bold; padding: 0;")
        self.messageLabel = QLabel(message, self)
        self.messageLabel.setStyleSheet(
            "font-size: 12px; font-weight: normal; padding: 0;")
        self.buttonClose = QPushButton(self)
        self.buttonClose.setIcon(QIcon(icon))
        self.buttonClose.setFlat(True)
        # self.buttonClose.setFixedSize(32, 32)
        self.buttonClose.setIconSize(QSize(30, 30))
        self.layout.addWidget(self.titleLabel)
        self.layout.addWidget(self.messageLabel, 2, 0)
        self.layout.addWidget(self.buttonClose, 0, 1)
        # self.setStyleSheet("border-radius: 100px;")
        # self.setStyleSheet("border :1px solid black;")


class Notification(QWidget):
    def __init__(self, parent=None):
        super(QWidget, self).__init__(parent=None)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        # self.setStyleSheet("background: #fdfdfd; padding: 0; border-radius:5px;")
        self.widget = QWidget(self)
        self.widget.setObjectName('Custom_Widget')
        # self.layout = QVBoxLayout(self)
        # self.layout.addWidget(self.widget)
        self.mainLayout = QVBoxLayout(self)

    def setNotify(self, title, message, icon, timeout):
        self.m = Message(title, message, icon)
        self.mainLayout.addWidget(self.m)
        self.m.buttonClose.clicked.connect(self.onClicked)
        self.show()
        QTimer.singleShot(timeout, 0, self.closeMe)

    def closeMe(self):
        self.close()
        self.m.close()

    def onClicked(self):
        self.close()

#####################################################


######## Dialog Zone ########

class Dialog(QWidget):

    NotChecked1 = []
    NotChecked2 = []
    result1 = {}
    result2 = {}
    file = ""

    def __init__(self):
        super().__init__()
        self.title = 'PyQt5 file dialogs - pythonspot.com'
        self.left = 10
        self.top = 10
        self.width = 640
        self.height = 480
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        self.openFileNameDialog()
        # output = self.openFileNamesDialog()
        # print("this is the output file", output)
        # self.saveFileDialog()

        self.show()

    def openFileNameDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(
            self, "فایل مورد نظر را انتخاب کنید", "", "All Files (*);;Python Files (*.py)", options=options)


        for i in Dialog.result2:
                # print(ref[i])
                Dialog.NotChecked2.append(ref[i])


class MainApp(QMainWindow, ui):

    record1 = 0
    record2 = 0

    def __init__(self, *args, **kwargs):
        QMainWindow.__init__(self)
        self.setupUi(self)

        date = jdatetime.datetime.now().strftime(" %Y-%m-%d")
        self.lbl_calender_2.setText(date)
        self.tabWidget.hide()
        self.Handle_Button()

 ################################ Handle Button ######################################
####################################################################################

    def Handle_Button(self):
        # pass
        self.btn_tafazol.clicked.connect(self.switchTab)
        self.btn_start.clicked.connect(self.showdialog)
        self.btn_person.clicked.connect(self.showNotification)
        self.btn_data.clicked.connect(self.showNotification)
        self.btn_manage.clicked.connect(self.showNotification)
        self.btn_dashboard.clicked.connect(self.tabWidget.hide)
        self.btn_call.clicked.connect(self.openNewWindow)
        self.btn_excel.clicked.connect(self.excel_out)

    def switchTab(self):
        self.tabWidget.show()
        self.tabWidget.setCurrentIndex(0)


################################ Excel Output ######################################
####################################################################################

    def excel_out(self):
        file = Dialog.file
        df = pd.read_csv(file)

        ref = df["ref"].copy().tolist()
        check = df["check"].copy().tolist()

        daftar = df["check"].copy().tolist()
        bank = df["ref"].copy().tolist()

        daftar_ = df["check"].copy().tolist()
        bank_ = df["ref"].copy().tolist()

        Exist = []
        for i, checkcell in enumerate(daftar):
            for j, refcell in enumerate(bank):
                # print(len(bank))
                if checkcell == refcell:
                    df['Barresi_Daftar_dar_Bank'][i] = 'Exist'
                    Exist.append(i)
                    bank.remove(refcell)
                #   bank[j] == "Checked"
                    break
                else:
                    df['Barresi_Daftar_dar_Bank'][i] = 'Not Exist'

        Exist = []
        for i, refcell in enumerate(bank_):
            for j, checkcell in enumerate(daftar_):
                # print(len(daftar_))
                if refcell == checkcell:
                    df['Barresi_Bank_dar_Daftar'][i] = 'Exist'
                    Exist.append(i)
                    daftar_.remove(checkcell)
                    break
                else:
                    df['Barresi_Bank_dar_Daftar'][i] = 'Not Exist'

        # for i, checkcell in enumerate(daftar):
        #     for j, refCell in enumerate(bank):
        #         if checkcell == refCell:
        #             df['Barresi_Daftar_dar_Bank'][i] = 'Exist'
        #             df['ref'][j] = "Checked"
        #             # df['output'][i].style.apply('background-color: green')
        #             break
        #         else:
        #             df['Barresi_Daftar_dar_Bank'][i] = 'Not Exist'

        # for i, checkcell in enumerate(bank):
        #     for j, refCell in enumerate(daftar):
        #         if checkcell == refCell:
        #             df['Barresi_Bank_dar_Daftar'][i] = 'Exist'
        #             df['check'][j] = "Checked"
        #             # df['output'][i].style.apply('background-color: green')
        #             break
        #         else:
        #             df['Barresi_Bank_dar_Daftar'][i] = 'Not Exist'

        list1 = df.index[df['Barresi_Daftar_dar_Bank'] == 'Not Exist'].tolist()
        list2 = df.index[df['Barresi_Bank_dar_Daftar'] == 'Not Exist'].tolist()

        # pdb.set_trace()

        df_ = pd.DataFrame()
        df_["Nmojood_Daftar"] = [check[i] for i in list1]

        df__ = pd.DataFrame()
        df__["Nmojood_Bank"] = [ref[i] for i in list2]

        df_concat = pd.concat([df, df_, df__], axis=1)

        # df["output"] = df["output"].map({'Exist': "True", 'Not Exist': "False"})
        outfile_name = (str(file).split('/')[-1]).split('.')[0] + ".xlsx"
        print(f"OutPut file name is: {outfile_name}")
        out_file = out_dir + outfile_name

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            sheet_name = "Bool"
            # Export DataFrame content
            df_concat.to_excel(writer, sheet_name=sheet_name)
            # Set backgrund colors depending on cell values
            sheet = writer.sheets[sheet_name]
            # Skip header row, process as many rows as there are DataFrames
            for cell, in sheet[f'D2:D{len(df_concat) + 1}']:
                # value is "True" or "False"
                value = df_concat["Barresi_Daftar_dar_Bank"].iloc[cell.row - 2]
                cell.fill = PatternFill("solid", start_color=(
                    "5cb800" if value == "Exist" else 'ff2800'))

            # Skip header row, process as many rows as there are DataFrames
            for cell, in sheet[f'E2:E{len(df_concat) + 1}']:
                # value is "True" or "False"
                value = df_concat["Barresi_Bank_dar_Daftar"].iloc[cell.row - 2]
                cell.fill = PatternFill("solid", start_color=(
                    "5cb800" if value == "Exist" else 'ff2800'))

        # reading the csv file
        df = pd.read_csv("Book2.csv")

        # updating the column value/data
        df.loc[:] = ''

        # writing into the file
        df.to_csv("Book2.csv", index=False)


################################ Notification ######################################
####################################################################################

    def showNotification(self, message):
        self.notification = Notification()
        message = {
            "ter": "شما امکان دسترسی به این بخش را ندارید. جهت فعالسازی با پشتیبانی تماس بگیرید."}
        if "done" in message.keys():
            icon = "message.png"
            self.notification.setNotify(
                "Successful", '%s' % message["done"], icon, 3000)
            r = QRect(self.x() + round(self.width() / 2) - round(self.notification.width() / 2),
                      self.y() + 26, self.notification.m.messageLabel.width() + 30, self.notification.m.messageLabel.height())
            self.notification.setGeometry(r)
        else:
            icon = "warning.png"
            self.notification.setNotify(
                "خطای دسترسی", '%s' % message["ter"], icon, 3000)
            r = QRect(self.x() + round(self.width() / 2) - round(self.notification.width() / 2),
                      self.y() + 26, self.notification.m.messageLabel.width() + 30, self.notification.m.messageLabel.height())
            self.notification.setGeometry(r)

################################ New Dialog ######################################
####################################################################################
    def openNewWindow(self):
        # Create a new dialog window
        # Create a new dialog window
        dialog = QDialog(self)
        dialog.setWindowTitle('تماس با ما')
        dialog.setFixedSize(350, 350)
        # Load the image

        image = QImage('./icons/logo.jpg')
        resizedImage = image.scaled(350, 350)

        # Create a label to display the image
        label = QLabel()
        label.setPixmap(QPixmap.fromImage(resizedImage))

        # Add the label to the dialog window
        layout = QVBoxLayout()
        layout.addWidget(label)
        dialog.setLayout(layout)

        # Set the dialog window's geometry
        dialog.setGeometry(100, 100, 300, 200)

        # Show the dialog window
        dialog.show()

##############################################################################################
    def showdialog(self):

        rec1 = int(self.line_recdaftar.text())
        rec2 = int(self.line_recbank.text())
        MainApp.record1 = rec1
        MainApp.record2 = rec2

        self.tableWidget.show()
        Dialog.NotChecked1 = []
        Dialog.NotChecked2 = []
        Dialog.result1 = {}
        Dialog.result2 = {}
        ex = Dialog()

        #############################################
        final_result1 = list(ex.result1)
        # print(final_result1)
        final_result1 = final_result1
        # print(final_result1)

        row_position_1 = len(ex.NotChecked2)
        row_position_2 = len(ex.NotChecked1)
        row_position = max(row_position_1, row_position_2)

        self.tableWidget.setRowCount(0)
        self.tableWidget.setRowCount(row_position)

        if final_result1:
            # self.tableWidget.setRowCount(0)
            # self.tableWidget.insertRow(0)
            for row, item in enumerate(ex.NotChecked1):
                # print(f"row:{row} and item:{item}")
                self.tableWidget.setItem(row, 0, QTableWidgetItem(str(item)))
                # for column, item in enumerate(form):
                #     self.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
                #     column += 1

                # row_position =self.tableWidget.rowCount()
                # self.tableWidget.insertRow(row_position)

        ##################################################

        final_result2 = list(ex.result2)
        # print(final_result2)
        final_result2 = final_result2
        # print(final_result2)

        if final_result2:
            # self.tableWidget.setRowCount(0)
            # self.tableWidget.insertRow(0)
            for row, item in enumerate(ex.NotChecked2):
                # print(f"row:{row} and item:{item}")
                self.tableWidget.setItem(row, 1, QTableWidgetItem(str(item)))
                # for column, item in enumerate(form):
                #     self.tableWidget.setItem(row, column, QTableWidgetItem(str(item)))
                #     column += 1


def main():
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    app.exec_()


if __name__ == "__main__":

    out_dir = './OutPut/'
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    ########## Check Security ############
    mac_address = uuid.getnode()
    mac_address = ':'.join(['{:02x}'.format(
        (uuid.getnode() >> ele) & 0xff) for ele in range(0, 8*6, 8)][::-1])
    main_hostname = "DESKTOP-1UQGEF3"
    hostname = socket.gethostname()
    main_mac_address = "88:d8:2e:72:a7:99"
    if main_hostname == hostname and main_mac_address == mac_address:
        # print(hostname)
        # print(mac_address)

        main()
